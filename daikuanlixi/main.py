import datetime
import math
from bs4 import BeautifulSoup
import openpyxl
import matplotlib.pyplot as plt
import numpy as np


class HousingLoan(object):
    def __init__(self, principal, benchmark_interest, floating_interest, years):
        """
            贷款总额：
            基准利率：
            浮动利率：
            贷款年限：
        """
        self.principal = principal
        self.benchmark_interest = benchmark_interest
        self.floating_interest = floating_interest
        # 年利率 = 基准利率(4.9) * 浮动利率 / 100(%)
        self.year_interest = self.benchmark_interest * self.floating_interest / 100.0
        # 月利率 = 年利率 / 12个月
        self.month_interest = self.year_interest / 12
        self.years = years
        # 贷款总期数 = 总年数 * 12个月
        self.total_period = self.years * 12


class EquivalentPrincipalAndInterest(HousingLoan):
    def __init__(self, principal, benchmark_interest, floating_interest, years):
        super().__init__(principal, benchmark_interest, floating_interest, years)

        self.month_repay = self.month_repay(self.principal, self.month_interest, self.total_period)
        self.repay_principal = []
        self.repay_accrual = []
        self.calculation(self.principal, self.month_interest, self.month_repay, self.repay_principal, self.repay_accrual)

    def month_repay(self, principal, month_interest, total_period):
        # 每月还款额 = (贷款本金*月利率* (1+月利率)^贷款月数) / ((1+月利率)^贷款月数-1)
        month_repayment = (principal * month_interest * (1 + month_interest) ** total_period) / \
                          ((1 + month_interest) ** total_period - 1)
        return np.array([month_repayment for i in total_period])

    def calculation(self, principal, month_interest, month_repay, repay_principal, repay_accrual):
        # 本月应还利息 = 贷款本金 * 月利率 * 1个月
        month_accrual = principal * month_interest * 1
        # 本月应还本金
        month_principal = month_repay - month_accrual

        # 添加本月的应还本金和应还利息
        self.repay_principal.append(month_principal)
        self.repay_accrual.append(month_accrual)

        # 每月余额
        balance = principal - month_principal
        # print('当月 应还本金:{} 应还利息:{} 余额:{} '.format(month_principal, month_accrual, balance))
        if balance >= 0.0001:
            self.calculation(balance, month_interest, month_repay, repay_principal, repay_accrual)


class EquivalentPrincipal(HousingLoan):
    def __init__(self, principal, benchmark_interest, floating_interest, years):
        super().__init__(principal, benchmark_interest, floating_interest, years)

        # # self.month_repay = []
        # self.repay_principal = self.principal/self.total_period
        # self.repay_accrual = []
        self.calculations = self.calculation(self.principal, self.total_period, self.month_interest)

    def calculation(self, principal, total_period, month_interest, repay_period=None):
        result = []
        # result.append(['应还本金', '应还利息', '月还款额', '已还本金', '已还利息', '贷款余额'])
        # 每月应还本金
        repay_principal = principal / total_period
        # 累计已还本金
        over_principal = 0
        over_accrual = 0
        # 计算还款几期，默认全部期数，也可以指定期数
        pre_period = total_period
        if repay_period:
            pre_period = repay_period

        for i in range(pre_period):
            # 创建一行列表
            rows = []

            # 每月利息=（本金-累计已还本金）×月利率
            month_accrual = (principal - over_principal) * month_interest

            # 累加已还本金和利息
            over_principal += repay_principal
            over_accrual += month_accrual

            rows.append(repay_principal)  # 应还本金
            rows.append(month_accrual)  # 应还利息
            rows.append(repay_principal + month_accrual)  # 月还款额
            rows.append(over_principal)  # 已还本金
            rows.append(over_accrual)  # 已还利息
            rows.append(principal - over_principal)  # 贷款余额
            # print(rows)
            result.append(rows)
        return np.array(result)

    def repay_money(self, start_date, repay_date, money=0):
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        repay_date = datetime.datetime.strptime(repay_date, "%Y-%m-%d")
        start_year = start_date.year
        repay_year = repay_date.year
        start_month = start_date.month
        repay_month = repay_date.month
        start_day = start_date.day
        repay_day = repay_date.day
        last_year = repay_year
        last_month = repay_month
        last_day = start_day
        new_day = last_day - repay_day
        if repay_day < start_day:
            # 求该月第一天
            repay_first_day = datetime.date(repay_year, repay_month, 1)
            # 求上月的最后一天
            pre_month = repay_first_day - datetime.timedelta(days=1)
            last_year = pre_month.year
            last_month = pre_month.month
            new_day = (pre_month.day - start_day) + repay_day
        print('最后一次还款日：{}年{}月{}日'.format(last_year, last_month, last_day))

        num = (last_year - start_year) * 12 + (last_month - start_month)
        result = self.calculation(self.principal, self.total_period, self.month_interest, repay_period=num)

        # 还款总额 = 提前还贷日前的最后一次还款后的剩余本金总额 + 提前还贷日前的最后一次还款日至提前还贷日止的利息
        # 提前还贷日前的最后一次还款日至提前还贷日止的利息 = 提前还贷日前的最后一次还款后的剩余本金总额 * （当期约定的贷款年利率 / 360） *提前还贷日前的最后一次还款日至提前还贷日止的天数

        # 提前还贷日前的最后一次还款后的剩余本金总额
        zonge = result[-1, -1]
        print(zonge)
        # 利息
        lixi = zonge * (self.year_interest / 360) * new_day
        # ['应还本金', '应还利息', '月还款额', '已还本金', '已还利息', '贷款余额']
        if money == 0:
            rows = np.array([[zonge, lixi, zonge+lixi, result[-1, 3]+zonge, result[-1, 4]+lixi, 0]])
            result = np.r_[result, rows]
            return result
        else:
            yue = zonge-money
            # print(yue)
            yunum = self.total_period - num
            rows = np.array([[money, lixi, money + lixi, result[-1, 3] + money, result[-1, 4] + lixi, yue]])
            post_result = self.calculation(yue, yunum, self.month_interest)
            # print(post_result)
            result = np.r_[result, rows]
            result = np.r_[result, post_result]
            return result


def excel():
    wb = openpyxl.load_workbook('tubiao.xlsx')
    ws = wb.active

    # 广发银行放贷计算器 http://www.cgbchina.com.cn/Channel/12122999
    with open('benxi.html', 'r', encoding='UTF-8') as f:
        html = f.read()

    soup = BeautifulSoup(html, 'lxml')
    rows = soup.find_all('tr')[1:]

    for row in rows:
            columns = row.find_all('td')
            i = int(columns[0].string)
            benjin = float(columns[1].string)
            lixi = float(columns[2].string)
            huankuan = float(columns[3].string)
            l_cnt = math.ceil(lixi/100)
            cnt = math.ceil(huankuan/100)
            for o in range(cnt):
                if o<= cnt - l_cnt:
                    ws.cell(row=i, column=o+1).style = 'Accent1'
                else:
                    ws.cell(row=i, column=o + 1).style = 'Accent2'

    with open('benjin.html', 'r', encoding='UTF-8') as f:
        html = f.read()

    soup = BeautifulSoup(html, 'lxml')
    rows = soup.find_all('tr')[1:]

    offset = 100
    for row in rows:
            columns = row.find_all('td')
            i = int(columns[0].string)
            benjin = float(columns[1].string)
            lixi = float(columns[2].string)
            huankuan = float(columns[3].string)
            l_cnt = math.ceil(lixi/100)
            cnt = math.ceil(huankuan/100)
            for o in range(cnt):
                if o <= cnt-l_cnt:
                    ws.cell(row=i, column=o+1+offset).style = 'Accent1'
                    # ws.cell(row=i, column=o + 1 + offset).style = 'Note'
                else:
                    ws.cell(row=i, column=o+1+offset).style = 'Accent2'

    for i in range(1,360):
        for o in range(1,250):
            if i%12==0 or o%10==0:
                ws.cell(row=i, column=o).style = 'Normal'
    # wb.save('tubiao.xlsx')
    # print('ok')


def main():
    # EPAI = EquivalentPrincipalAndInterest(1620000, 4.9, 1.2, 30)
    # print(EPAI.repay_accrual)
    EP = EquivalentPrincipal(1620000, 4.9, 1.2, 30)
    # result = EP.calculations
    result = EP.repay_money('2019-02-15', '2021-01-20', 500000)

    repay_principal = result[0:result.shape[0], 0]
    repay_accrual = result[0:result.shape[0], 1]
    index = np.arange(len(repay_principal))
    plt.barh(index, repay_principal, 1, alpha=0.5)
    plt.barh(index, repay_accrual, 1, left=repay_principal, color='red', alpha=0.5)

    result = EP.calculations

    repay_principal = result[0:result.shape[0], 0]
    repay_accrual = result[0:result.shape[0], 1]
    index = np.arange(len(repay_principal))
    plt.barh(index, repay_principal, 1, alpha=0.5)
    plt.barh(index, repay_accrual, 1, left=repay_principal, color='red', alpha=0.5)

    # 中文乱码的处理
    # plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
    # plt.rcParams['axes.unicode_minus'] = False

    # plt.xlabel('金额')
    # ticks = ['第{}期'.format(i) for i in range(1, 361) if (i % 1) == 0]
    # plt.yticks(range(len(ticks)), ticks)
    # plt.xlim([0, 15000])
    # plt.title("等额本金还款比例图", fontproperties="SimHei")  # （黑体）


    plt.show()


if __name__ == '__main__':
    main()

